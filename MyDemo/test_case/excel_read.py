'''
    读取Excel文件
    基于Excel文件的内容去进行读取，并结合获取的数据进行自动化的测试执行
'''

import openpyxl

from MyDemo.web_keys.keys import Keys

def read(file):
    # 读取Excel表格中的用例正文
    excel = openpyxl.load_workbook(file)
    # 获取页签，遍历页签中的所有内容
    for name in excel.sheetnames:
        print(excel.sheetnames)
        sheet = excel[name]
        print('***************{}***************'.format(name))
        for values in sheet.values:
            # print(type(values))
            # print(values)
            '''
                所得结果如下：
                ***************Sheet1***************
                ('测试用例', None, None, None, None, None)
                ('编号', '操作行为', '参数1(定位方法', '参数2(定位值', '参数3(文本', '描述')
                (1, 'open_browser', None, None, 'Chrome', '启动浏览器')
                (2, 'open', None, None, 'http://www.baidu.com', '访问url')
                (3, 'input', 'id', 'kw', '虚竹', '输入文本')
                (4, 'click', 'id', 'su', None, '点击搜索')
                (5, 'wait', None, None, 5, '强制等待5秒')
                (6, 'quit', None, None, None, '退出浏览器')
            '''
            # 根据获取的数据内容我们知道，如果第一个单元格为int类型，说明测试用例正文进入测试用例数据内容
            if type(values[0]) is int:
                print(values)
                # 正常获取数据后，进行数据拆分，使用一个data字典对读取的数据进行接收，便于后续操作
                data = {}
                data['name'] = values[2]
                data['value'] = values[3]
                data['txt'] = values[4]
                data['expect'] = values[6]
                print(data)
                '''
                所得结果如下：
                (1, 'open_browser', None, None, 'Chrome', '启动浏览器')
                {'name': None, 'value': None, 'txt': 'Chrome'}
                '''
                # 清除参数字典中为None的参数键值对，将字典中的键放进list中进行遍历，再每一个去访问，若为None则删除该键值对
                for key in list(data.keys()):
                    if data[key] is None:
                        del data[key]
                print(data)

                # 基于操作行为和对应参数来执行自动化操作（1.对象实例化,2.实例化后常规操作行为）
                '''
                    用例的操作行为主要分为：
                        1. 实例化，通过一个操作行为实例化关键字驱动类对象
                        2. 常规操作，通过调用已实例化的对象，执行对应的函数。
                        3. 断言操作，判断预期与实际是否符合，将结果填入测试用例中。
                '''
                # 实例化操作
                if values[1] == 'open_browser':
                    # driver = Keys(values['txt'])
                    keys = Keys(**data)
                elif 'assert' in values[1]:
                    status = getattr(keys, values[1])(**data)
                    if status:
                        sheet.cell(row=values[0] + 2,column=8).value = 'Pass'
                    else:
                        sheet.cell(row=values[0] + 2,column=8).value = 'Failed'
                # 常规操作
                else:
                    # 实例化对象.方法名，实例化对象已经再if中创建（必有）
                    getattr(keys, values[1])(**data)
    excel.save(file)
    print('保存成功')
